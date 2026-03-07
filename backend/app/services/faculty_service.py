"""
Faculty service — backed by SQLite (file metadata) + local filesystem (file bytes).

File flow:
  Upload  → bytes written to  backend/uploads/<uuid>_<filename>
            metadata row inserted into uploaded_files table
  Parse   → bytes read from   file_path on disk
  Delete  → file_path unlinked from disk + row deleted from DB
"""

import csv
import io
import json
import logging
import os
import re
import statistics
import uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import inspect

import app.core.database as _db
from app.core.database import UPLOAD_DIR
from app.models.uploaded_file import UploadedFile
from app.models.student_mark import StudentMark
from app.core import data_store as db  # only for filter constants (DEPARTMENTS etc.)


logger = logging.getLogger(__name__)

# Grade-bucket colours (A → F)
_GRADE_COLORS = ["#6366f1", "#8b5cf6", "#a78bfa", "#c4b5fd", "#ddd6fe"]

# Column name alias sets
_NAME_ALIASES   = {"name", "student", "student_name", "student name", "sname", "studentname"}
_ROLL_ALIASES_HIGH = {
    "roll", "roll_no", "rollno", "roll no", "roll number", "rollnumber",
    "reg", "reg_no", "regno", "student_id",
}
_ROLL_ALIASES_LOW = {"sno", "s.no", "serial", "id", "no", "sl.no", "sl no", "sr.no", "sr no"}
_ROLL_ALIASES = _ROLL_ALIASES_HIGH | _ROLL_ALIASES_LOW
_MARKS_ALIASES = {
    "marks", "score", "scores", "total", "total_marks", "totalmarks",
    "final", "final_marks", "finalmarks", "grade_points", "obtained",
    "marks_obtained", "result", "exam", "exam_marks", "marks/100",
    "grand total", "grandtotal", "overall", "aggregate", "sum",
}


# ── Column helpers ────────────────────────────────────────────────────────────

def _clean_header(h: str) -> str:
    s = str(h).strip().lower()
    if "(" in s:
        s = s[:s.index("(")].strip()
    s = s.rstrip(" 0123456789")
    return s.strip()

def _is_name_col(h: str)      -> bool: return _clean_header(h) in _NAME_ALIASES or str(h).strip().lower() in _NAME_ALIASES
def _is_roll_col(h: str)      -> bool: return str(h).strip().lower() in _ROLL_ALIASES
def _is_roll_col_high(h: str) -> bool: return str(h).strip().lower() in _ROLL_ALIASES_HIGH
def _is_marks_col(h: str)     -> bool:
    clean = _clean_header(h)
    return clean in _MARKS_ALIASES or str(h).strip().lower() in _MARKS_ALIASES

def _is_serial_like_col(h: str) -> bool:
    """Detect S.No/Serial columns so they don't appear as marks components."""
    return str(h).strip().lower() in _ROLL_ALIASES_LOW or _clean_header(h) in _ROLL_ALIASES_LOW

def _looks_like_serial_col(col: str, rows: list[dict]) -> bool:
    """
    Detect serial-number columns even when the header is blank/unknown.

    Typical examples are 1..N, 0..N-1, mostly unique, small integers.
    """
    values: list[int] = []
    for r in rows[:30]:
        v = r.get(col)
        n = _try_parse_number(v)
        if n is None:
            continue
        if abs(n - round(n)) > 1e-9:
            return False
        values.append(int(round(n)))
        if len(values) >= 12:
            break
    if len(values) < 6:
        return False
    uniq = set(values)
    if len(uniq) / len(values) < 0.9:
        return False
    mn, mx = min(values), max(values)
    # reject obviously-not-serial columns
    if mx > 200:
        return False
    # common serial forms
    if mn in (0, 1) and mx <= len(uniq) + 3:
        return True
    return False

def _detect_header_row_index(rows_raw: list[tuple]) -> int:
    """
    XLSX sheets sometimes have a title row above the real header.
    Pick the first row with the best "header-like" score.
    """
    best_i = 0
    best_score = -1
    for i in range(min(10, len(rows_raw))):
        row = rows_raw[i]
        score = 0
        non_empty = 0
        for cell in row:
            if cell is None or str(cell).strip() == "":
                continue
            non_empty += 1
            text = str(cell).strip()
            if _is_name_col(text):
                score += 3
            elif _is_roll_col(text) or _is_roll_col_high(text):
                score += 3
            elif _is_marks_col(text):
                score += 3
            else:
                clean = _clean_header(text)
                if any(t in clean.replace(" ", "") for t in ("assign", "bit", "quiz", "mid", "unit", "internal", "external", "descriptive", "lab", "project", "viva")):
                    score += 1
        # ignore almost-empty rows (common for titles)
        if non_empty <= 1:
            continue
        if score > best_score:
            best_score = score
            best_i = i
    # Require at least 2 strong hits; otherwise keep the first row.
    return best_i if best_score >= 6 else 0

def _dedupe_total_like_columns(columns: list[str]) -> list[str]:
    """
    Avoid duplicate Total/Marks columns (e.g. 'Total', 'Total (100)', 'TOTAL ').
    Keep the first total-like column in the list and drop the rest.
    """
    out: list[str] = []
    total_seen = False
    for c in columns:
        if _is_marks_col(c):
            if total_seen:
                continue
            total_seen = True
        out.append(c)
    return out

def _is_numeric_col(col: str, rows: list[dict]) -> bool:
    vals = [rows[i].get(col) for i in range(min(20, len(rows)))]
    numeric = total = 0
    for v in vals:
        if v is None or str(v).strip() == "": continue
        total += 1
        try:
            float(str(v).replace(",", ""))
            numeric += 1
        except ValueError:
            pass
    return total > 0 and (numeric / total) >= 0.7

def _col_mean(col: str, rows: list[dict]) -> float:
    vals = []
    for r in rows[:50]:
        v = r.get(col)
        if v is None: continue
        try:
            vals.append(float(str(v).replace(",", "")))
        except (TypeError, ValueError):
            pass
    return sum(vals) / len(vals) if vals else 0.0


_NUM_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")


def _try_parse_number(value) -> float | None:
    """
    Parse numeric-like cell values.

    Supports formats like:
    - 10, 10.5, "10", "10.0"
    - "9/10" (takes the first number)
    - "10 (out of 10)" (takes first number)
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    s_low = s.lower()
    if s_low in {"absent", "ab", "a", "na", "n/a", "none", "nan", "-"}:
        return None
    # Try fast float
    try:
        return float(s.replace(",", ""))
    except Exception:
        pass
    # Extract first number
    m = _NUM_RE.search(s.replace(",", ""))
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _col_numeric_count(col: str, rows: list[dict], limit: int = 50) -> int:
    count = 0
    for r in rows[:limit]:
        v = r.get(col)
        if _try_parse_number(v) is not None:
            count += 1
    return count


def _normalize_roll_no(value) -> str:
    """
    Normalize roll numbers coming from CSV/XLSX cells.

    Common issues:
    - Excel stores numbers as floats: 2301 -> 2301.0
    - Leading/trailing whitespace
    """
    if value is None:
        return ""

    # Preserve non-numeric IDs (e.g. "CS2023045") as strings.
    if isinstance(value, (int,)):
        return str(value).strip()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


# ── File parsing ──────────────────────────────────────────────────────────────

def _student_marks_has_components(session) -> bool:
    try:
        insp = inspect(session.get_bind())
        cols = {c["name"] for c in insp.get_columns("student_marks")}
        return "components_json" in cols
    except Exception:
        return False

def _read_file_bytes(file_path: str) -> bytes:
    try:
        return Path(file_path).read_bytes()
    except Exception as e:
        logger.warning("Cannot read file %s: %s", file_path, e)
        return b""


def _parse_marks_from_path(file_path: str, filename: str) -> list[dict]:
    raw = _read_file_bytes(file_path)
    if not raw:
        return []
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return _parse_xlsx(raw)
    if ext == "pdf":
        return _parse_pdf(raw)
    return _parse_csv(raw)


def _parse_csv(raw: bytes) -> list[dict]:
    try:
        text = raw.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            reader2 = csv.DictReader(io.StringIO(text), delimiter=";")
            rows = list(reader2)
    except Exception as e:
        logger.warning("CSV parse error: %s", e)
        return []
    return _extract_marks(rows)


def _parse_xlsx(raw: bytes) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.warning("XLSX parse error: %s", e)
        return []
    if not rows_raw:
        return []
    header_idx = _detect_header_row_index(rows_raw)
    header_row = rows_raw[header_idx] if header_idx < len(rows_raw) else rows_raw[0]
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    rows = []
    for row in rows_raw[header_idx + 1 :]:
        if any(c is not None and str(c).strip() != "" for c in row):
            rows.append(dict(zip(headers, row)))
    return _extract_marks(rows)


def _parse_pdf(raw: bytes) -> list[dict]:
    try:
        import pdfplumber
        rows: list[dict] = []
        headers: list[str] = []
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                if not headers:
                    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(table[0])]
                    for data_row in table[1:]:
                        if any(c and str(c).strip() for c in data_row):
                            rows.append(dict(zip(headers, data_row)))
                else:
                    for data_row in table[1:]:
                        if any(c and str(c).strip() for c in data_row):
                            rows.append(dict(zip(headers, data_row)))
    except ImportError:
        logger.warning("pdfplumber not installed")
        return []
    except Exception as e:
        logger.warning("PDF parse error: %s", e)
        return []
    return _extract_marks(rows)


def _extract_marks(rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    headers = list(rows[0].keys())

    name_col = next((c for c in headers if _is_name_col(c)), None)
    if name_col is None:
        name_col = next(
            (c for c in headers if not _is_roll_col(c) and not _is_numeric_col(c, rows)),
            headers[0],
        )

    roll_col = next((c for c in headers if _is_roll_col_high(c)), None)
    if roll_col is None:
        roll_col = next((c for c in headers if _is_roll_col(c)), None)

    excluded: set[str] = {c for c in (name_col, roll_col) if c is not None}
    # Exclude serial-number columns so they don't become components (even if the header is blank).
    excluded |= {c for c in headers if _is_serial_like_col(c) or _looks_like_serial_col(c, rows)}

    marks_col = next((c for c in headers if _is_marks_col(c) and c not in excluded), None)
    if marks_col is None:
        numeric_candidates = [c for c in headers if c not in excluded and _is_numeric_col(c, rows)]
        if numeric_candidates:
            marks_col = max(numeric_candidates, key=lambda c: _col_mean(c, rows))

    if not marks_col:
        logger.warning("No marks column found. headers=%s", headers)
        return []

    def _is_component_header(h: str) -> bool:
        clean = _clean_header(h)
        if not clean:
            return False
        # exclude obvious non-score columns
        if clean in {"grade", "result", "remarks", "remark", "status"}:
            return False
        tokens = {
            "assign", "assignment", "ass", "bit", "bits", "quiz", "mid", "midterm", "unit", "test",
            "internal", "external", "theory", "practical", "lab", "project", "viva", "descriptive",
        }
        return any(t in clean.replace(" ", "") for t in tokens)

    # Component columns: include any column that has at least one numeric-like value in the sample,
    # OR the header looks like a marks component (so sparse columns still show up).
    metric_cols = [c for c in headers if c not in excluded and (_col_numeric_count(c, rows) > 0 or _is_component_header(c))]
    if marks_col and marks_col not in metric_cols:
        metric_cols.append(marks_col)

    result = []
    for r in rows[:500]:
        try:
            raw_name = str(r.get(name_col, "")).strip()
            if not raw_name or raw_name.lower() in ("none", "nan", ""):
                continue

            components: dict[str, float | None] = {}
            for c in metric_cols:
                components[c] = _try_parse_number(r.get(c))

            raw_val = r.get(marks_col)
            if raw_val is None:
                other = [
                    c for c in headers
                    if c not in excluded and c != marks_col
                    and not _is_roll_col(c) and _is_numeric_col(c, rows)
                ]
                computed, ok = 0.0, False
                for c in other:
                    v = r.get(c)
                    if v is not None:
                        parsed = _try_parse_number(v)
                        if parsed is not None:
                            computed += parsed
                            ok = True
                if not ok:
                    continue
                marks = computed
            else:
                parsed = _try_parse_number(raw_val)
                if parsed is None:
                    continue
                marks = parsed

            roll_val = ""
            if roll_col:
                roll_val = _normalize_roll_no(r.get(roll_col, ""))
                if roll_val.isdigit() and not _is_roll_col_high(roll_col):
                    better = next(
                        (c for c in headers
                         if c not in excluded and c != name_col and c != marks_col
                         and not _is_numeric_col(c, rows) and c != roll_col),
                        None,
                    )
                    if better:
                        roll_val = _normalize_roll_no(r.get(better, ""))

            result.append({"name": raw_name, "roll_no": roll_val, "marks": marks, "components": components})
        except (TypeError, ValueError):
            continue
    return result


def _parse_raw_rows(file_path: str, filename: str) -> tuple[list[dict], list[str]]:
    raw = _read_file_bytes(file_path)
    if not raw:
        return [], []
    try:
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw:
                return [], []
            header_idx = _detect_header_row_index(rows_raw)
            header_row = rows_raw[header_idx] if header_idx < len(rows_raw) else rows_raw[0]
            headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
            rows = [dict(zip(headers, row)) for row in rows_raw[header_idx + 1 :] if any(c is not None for c in row)]
        else:
            text = raw.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            headers = list(reader.fieldnames or [])
        return rows, headers
    except Exception as e:
        logger.warning("_parse_raw_rows error: %s", e)
        return [], []


# ── DB helpers ────────────────────────────────────────────────────────────────

def _all_files(faculty_user_id: str) -> list[dict]:
    with _db.SessionLocal() as session:
        files = (
            session.query(UploadedFile)
            .filter(UploadedFile.uploaded_by_user_id == faculty_user_id)
            .order_by(UploadedFile.created_at.desc())
            .all()
        )
        return [f.to_dict() for f in files]


def _file_record(file_id: str, faculty_user_id: str) -> UploadedFile | None:
    with _db.SessionLocal() as session:
        return (
            session.query(UploadedFile)
            .filter(
                UploadedFile.id == file_id,
                UploadedFile.uploaded_by_user_id == faculty_user_id,
            )
            .first()
        )


def _parse_marks(file_id: str) -> list[dict]:
    with _db.SessionLocal() as session:
        rec = session.get(UploadedFile, file_id)
        if not rec:
            return []
        return _parse_marks_from_path(rec.file_path, rec.name)


# ── Filtering helpers ─────────────────────────────────────────────────────────

def _filter_files(
    faculty_user_id: str,
    department: str | None = None,
    year: str | None = None,
    section: str | None = None,
    subject: str | None = None,
) -> list[dict]:
    files = _all_files(faculty_user_id)
    result = []
    for f in files:
        if department is not None and department != "" and (f.get("department") or "") != department: continue
        if year is not None and year != "" and (f.get("year") or "") != year: continue
        if section is not None and section != "" and (f.get("section") or "") != section: continue
        if subject is not None and subject != "" and (f.get("subject") or "") != subject: continue
        result.append(f)
    return result


def _marks_from_files(files: list[dict]) -> list[dict]:
    result: list[dict] = []
    for f in files:
        result.extend(_parse_marks(f["id"]))
    return result


def _marks_from_file_ids(file_ids: list[str]) -> list[dict]:
    if not file_ids:
        return []
    with _db.SessionLocal() as session:
        rows = (
            session.query(StudentMark.student_name, StudentMark.roll_no, StudentMark.marks)
            .filter(StudentMark.uploaded_file_id.in_(file_ids))
            .all()
        )
        return [
            {"name": name, "roll_no": (roll_no or ""), "marks": float(marks)}
            for (name, roll_no, marks) in rows
        ]


# ── Grade distribution ────────────────────────────────────────────────────────

def _grade_distribution(marks: list[float]) -> list[dict]:
    buckets: dict[str, int] = {
        "A (≥85)": 0, "B (70-84)": 0, "C (55-69)": 0, "D (40-54)": 0, "F (<40)": 0,
    }
    for v in marks:
        if v >= 85:   buckets["A (≥85)"] += 1
        elif v >= 70: buckets["B (70-84)"] += 1
        elif v >= 55: buckets["C (55-69)"] += 1
        elif v >= 40: buckets["D (40-54)"] += 1
        else:         buckets["F (<40)"] += 1
    return [
        {"name": k, "value": v, "color": c}
        for (k, v), c in zip(buckets.items(), _GRADE_COLORS) if v > 0
    ]


def _col_stats_multi(rows: list[dict], score_cols: list[str]) -> list[dict]:
    result = []
    for col in score_cols:
        vals = []
        for r in rows:
            v = r.get(col)
            if v is None: continue
            try:
                vals.append(float(str(v).replace(",", "")))
            except (TypeError, ValueError):
                pass
        if not vals: continue
        result.append({
            "name":   str(col),
            "mean":   round(statistics.mean(vals), 2),
            "median": round(statistics.median(vals), 2),
            "min":    min(vals),
            "max":    max(vals),
            "stdev":  round(statistics.stdev(vals), 2) if len(vals) > 1 else 0.0,
        })
    return result


# ── Public API ────────────────────────────────────────────────────────────────

def get_stats(
    faculty_user_id: str,
    department: str | None = None,
    year: str | None = None,
    section: str | None = None,
    subject: str | None = None,
) -> dict:
    files = _filter_files(faculty_user_id, department, year, section, subject)
    all_marks = _marks_from_file_ids([f["id"] for f in files])
    total_docs = len(_all_files(faculty_user_id))

    if not all_marks:
        return {
            "total_students": 0, "total_students_change": "—",
            "avg_performance": 0.0, "avg_performance_change": "—",
            "total_documents": total_docs, "total_documents_change": "—",
            "pass_rate": 0.0, "pass_rate_change": "—",
        }
    marks_vals = [m["marks"] for m in all_marks]
    return {
        "total_students": len(all_marks), "total_students_change": "—",
        "avg_performance": round(statistics.mean(marks_vals), 1), "avg_performance_change": "—",
        "total_documents": total_docs, "total_documents_change": "—",
        "pass_rate": round(100 * sum(1 for v in marks_vals if v >= 40) / len(marks_vals), 1),
        "pass_rate_change": "—",
    }


def get_files(faculty_user_id: str) -> list[dict]:
    return _all_files(faculty_user_id)


def get_file_by_id(file_id: str, faculty_user_id: str) -> dict:
    with _db.SessionLocal() as session:
        rec = (
            session.query(UploadedFile)
            .filter(
                UploadedFile.id == file_id,
                UploadedFile.uploaded_by_user_id == faculty_user_id,
            )
            .first()
        )
        if not rec:
            raise HTTPException(status_code=404, detail=f"File '{file_id}' not found.")
        return rec.to_dict()


def add_file(
    upload: UploadFile,
    subject: str,
    department: str = "",
    year: str = "",
    section: str = "",
    faculty_user_id: str = "",
) -> dict:
    content = upload.file.read()
    size_bytes = len(content)
    size_label = (
        f"{size_bytes / (1024 * 1024):.1f} MB" if size_bytes >= 1_048_576
        else f"{size_bytes // 1024} KB" if size_bytes >= 1024
        else f"{size_bytes} B"
    )

    fname = upload.filename or f"upload_{uuid.uuid4().hex[:8]}.csv"
    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext not in {"csv", "xlsx", "xls", "pdf"}:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '.{ext}'.")

    file_id  = str(uuid.uuid4())
    safe_name = f"{file_id}_{fname}"
    file_path = str(UPLOAD_DIR / safe_name)

    # Write bytes to disk
    Path(file_path).write_bytes(content)

    # Insert metadata + parsed marks into DB (single commit)
    with _db.SessionLocal() as session:
        try:
            now = datetime.now(timezone.utc)
            rec = UploadedFile(
                id=file_id,
                name=fname,
                date=datetime.now().strftime("%b %d, %Y"),
                subject=(subject or "General").strip() or "General",
                department=(department or "").strip(),
                year=(year or "").strip(),
                section=(section or "").strip(),
                size=size_label,
                file_path=file_path,
                uploaded_by_user_id=faculty_user_id or None,
                created_at=now,
            )
            session.add(rec)
            # Ensure the parent row exists before inserting dependent marks rows.
            # Without this explicit flush, SQLAlchemy may batch INSERTs in an order
            # that violates the FK constraint on Postgres.
            session.flush()

            marks_rows = _parse_marks_from_path(file_path, fname)
            if not marks_rows:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "Could not extract student marks. Ensure the file has "
                        "'name'/'student' and 'marks'/'score'/'total' columns."
                    ),
                )

            marks_vals = [m["marks"] for m in marks_rows]
            max_mark = max(marks_vals) if marks_vals else 0
            if max_mark > 100:
                scale = 100.0 / max_mark
                for i, m in enumerate(marks_rows):
                    m["marks"] = round(marks_vals[i] * scale, 1)

            has_components = _student_marks_has_components(session)
            mark_objs: list[StudentMark] = []
            for m in marks_rows:
                student_name = (m.get("name") or "").strip() or "Unknown"
                roll_no = _normalize_roll_no(m.get("roll_no") or "") or None
                components = m.get("components") or {}
                components_json = json.dumps(components, ensure_ascii=False) if (has_components and components) else None
                mark_objs.append(
                    StudentMark(
                        id=str(uuid.uuid4()),
                        uploaded_file_id=file_id,
                        student_name=student_name,
                        roll_no=roll_no,
                        marks=float(m["marks"]),
                        components_json=components_json,
                        created_at=now,
                    )
                )

            session.add_all(mark_objs)
            session.commit()
            return rec.to_dict()
        except Exception:
            session.rollback()
            # Don't leave dead files on disk if DB insert fails / parse fails.
            try:
                os.unlink(file_path)
            except FileNotFoundError:
                pass
            raise


def delete_file(file_id: str, faculty_user_id: str) -> dict:
    with _db.SessionLocal() as session:
        rec = (
            session.query(UploadedFile)
            .filter(
                UploadedFile.id == file_id,
                UploadedFile.uploaded_by_user_id == faculty_user_id,
            )
            .first()
        )
        if not rec:
            raise HTTPException(status_code=404, detail=f"File '{file_id}' not found.")
        name = rec.name
        # Delete from disk
        try:
            os.unlink(rec.file_path)
        except FileNotFoundError:
            pass
        # Delete derived marks (SQLite FK cascade isn't guaranteed unless PRAGMA is enabled).
        session.query(StudentMark).filter(StudentMark.uploaded_file_id == file_id).delete(
            synchronize_session=False
        )
        session.delete(rec)
        session.commit()
    return {"message": f"File '{name}' deleted successfully."}


def analyze_file(file_id: str, faculty_user_id: str) -> dict:
    with _db.SessionLocal() as session:
        rec = (
            session.query(UploadedFile)
            .filter(
                UploadedFile.id == file_id,
                UploadedFile.uploaded_by_user_id == faculty_user_id,
            )
            .first()
        )
        if not rec:
            raise HTTPException(status_code=404, detail=f"File '{file_id}' not found.")
        file_meta = rec.to_dict()
        file_path = rec.file_path
        fname = rec.name

    raw = _read_file_bytes(file_path)
    if not raw:
        raise HTTPException(status_code=422, detail="No content found. Please re-upload the file.")

    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext in ("xlsx", "xls"):
        student_marks = _parse_xlsx(raw)
    elif ext == "pdf":
        student_marks = _parse_pdf(raw)
    else:
        student_marks = _parse_csv(raw)

    if not student_marks:
        raise HTTPException(
            status_code=422,
            detail="Could not extract student marks. Ensure the file has 'name'/'student' and 'marks'/'score'/'total' columns.",
        )

    marks_vals = [m["marks"] for m in student_marks]
    max_mark = max(marks_vals)
    if max_mark > 100:
        scale = 100.0 / max_mark
        marks_vals = [round(v * scale, 1) for v in marks_vals]
        for i, m in enumerate(student_marks):
            m["marks"] = marks_vals[i]

    grade_dist = _grade_distribution(marks_vals)
    all_rows, headers = _parse_raw_rows(file_path, fname)

    score_cols = [
        h for h in headers
        if not _is_name_col(h) and not _is_roll_col(h) and _is_numeric_col(h, all_rows)
    ] if all_rows else []

    col_stats = _col_stats_multi(all_rows, score_cols) if score_cols else [{
        "name": "marks",
        "mean": round(statistics.mean(marks_vals), 2),
        "median": round(statistics.median(marks_vals), 2),
        "min": min(marks_vals),
        "max": max(marks_vals),
        "stdev": round(statistics.stdev(marks_vals), 2) if len(marks_vals) > 1 else 0.0,
    }]

    from app.services import ml_service
    is_multi = len(score_cols) >= 2 and len(all_rows) >= 3
    ml_result = ml_service.predict_multi(student_marks, all_rows, headers) if is_multi else ml_service.predict(student_marks)

    predicted_vs_actual = []
    if ml_result.get("lr_available"):
        for p in ml_result["predictions"]:
            if p.get("predicted_marks") is not None:
                predicted_vs_actual.append({
                    "name": p["name"], "actual": p["marks"], "predicted": p["predicted_marks"],
                })

    return {
        "file_id":            file_id,
        "file_name":          file_meta["name"],
        "subject":            file_meta["subject"],
        "department":         file_meta.get("department", ""),
        "year":               file_meta.get("year", ""),
        "section":            file_meta.get("section", ""),
        "row_count":          len(student_marks),
        "columns":            col_stats,
        "grade_distribution": grade_dist,
        "student_marks":      [{"name": m["name"], "marks": int(m["marks"])} for m in student_marks],
        "ml_predictions":     ml_result["predictions"],
        "class_insights":     ml_result["class_insights"],
        "lr_available":       ml_result.get("lr_available", False),
        "has_multi_column":   ml_result.get("has_multi_column", False),
        "predicted_vs_actual": predicted_vs_actual,
    }


def get_analytics(faculty_user_id: str, department, year, section, subject) -> dict:
    files = _filter_files(faculty_user_id, department, year, section, subject)
    all_marks = _marks_from_file_ids([f["id"] for f in files])
    if not all_marks:
        return {
            "student_marks": [], "performance_trend": [], "grade_distribution": [],
            "section_breakdown": [], "student_detail_list": [],
            "filters_applied": {"department": department, "year": year, "section": section, "subject": subject},
        }
    marks_vals = [m["marks"] for m in all_marks]
    avg = round(statistics.mean(marks_vals), 1)
    grade_dist = _grade_distribution(marks_vals)
    section_breakdown = _compute_section_breakdown(faculty_user_id, department, year, subject)
    from app.services import ml_service
    ml_result = ml_service.predict(all_marks)
    return {
        "student_marks":      [{"name": m["name"], "marks": int(m["marks"])} for m in all_marks],
        "performance_trend":  [{"month": "Current", "avg": avg}],
        "grade_distribution": grade_dist,
        "section_breakdown":  section_breakdown,
        "student_detail_list": ml_result["predictions"],
        "filters_applied": {"department": department, "year": year, "section": section, "subject": subject},
    }


def _compute_section_breakdown(faculty_user_id: str, department, year, subject) -> list[dict]:
    section_map: dict[str, list[float]] = {}
    with _db.SessionLocal() as session:
        q = (
            session.query(UploadedFile.section, StudentMark.marks)
            .join(StudentMark, StudentMark.uploaded_file_id == UploadedFile.id)
            .filter(UploadedFile.uploaded_by_user_id == faculty_user_id)
        )
        if department:
            q = q.filter(UploadedFile.department == department)
        if year:
            q = q.filter(UploadedFile.year == year)
        if subject:
            q = q.filter(UploadedFile.subject == subject)

        for sec, marks in q.all():
            sec_label = sec or "Unknown"
            section_map.setdefault(sec_label, []).append(float(marks))
    result = []
    for sec, vals in sorted(section_map.items()):
        avg = round(statistics.mean(vals), 1)
        pass_rate = round(100 * sum(1 for v in vals if v >= 40) / len(vals), 1)
        result.append({"section": sec, "avg": avg, "pass_rate": pass_rate, "total_students": len(vals)})
    return result


def generate_average_report(file_ids: list[str], faculty_user_id: str) -> dict:
    if len(file_ids) < 2:
        raise HTTPException(status_code=400, detail="At least 2 files must be selected.")
    # Validate ownership up-front (avoid leaking existence across faculty accounts).
    for fid in file_ids:
        get_file_by_id(fid, faculty_user_id)
    combined = _marks_from_file_ids(file_ids)
    if not combined:
        raise HTTPException(status_code=422, detail="No parseable data found in the selected files.")
    marks_vals = [m["marks"] for m in combined]
    return {
        "avg_score":          round(statistics.mean(marks_vals), 1),
        "pass_rate":          round(100 * sum(1 for v in marks_vals if v >= 40) / len(marks_vals), 1),
        "highest_score":      max(marks_vals),
        "lowest_score":       min(marks_vals),
        "student_marks":      combined[:20],
        "grade_distribution": _grade_distribution(marks_vals),
        "source_files":       file_ids,
    }


def get_filter_options() -> dict:
    return {
        "departments": db.DEPARTMENTS,
        "years":       db.YEARS,
        "sections":    db.SECTIONS,
        "subjects":    db.SUBJECTS,
    }


def get_student_list(faculty_user_id: str, file_ids: list[str] | None = None) -> list[dict]:
    """
    Return a simple student list (name/roll/marks) for the faculty's uploaded files.

    If file_ids is provided, it is scoped to those uploads (ownership enforced).
    """
    if file_ids:
        for fid in file_ids:
            get_file_by_id(fid, faculty_user_id)

    with _db.SessionLocal() as session:
        q = (
            session.query(
                StudentMark.student_name,
                StudentMark.roll_no,
                StudentMark.marks,
                UploadedFile.subject,
                UploadedFile.id,
            )
            .join(UploadedFile, UploadedFile.id == StudentMark.uploaded_file_id)
            .filter(UploadedFile.uploaded_by_user_id == faculty_user_id)
        )
        if file_ids:
            q = q.filter(StudentMark.uploaded_file_id.in_(file_ids))

        rows = q.order_by(UploadedFile.created_at.desc()).all()
        return [
            {
                "file_id": file_id,
                "subject": subject or "Unknown",
                "name": (student_name or "").strip() or "Unknown",
                "roll_no": (roll_no or "").strip(),
                "marks": float(marks),
            }
            for (student_name, roll_no, marks, subject, file_id) in rows
        ]


def _loads_components(value: str | None) -> dict[str, float | None]:
    if not value:
        return {}
    try:
        obj = json.loads(value)
    except Exception:
        return {}
    if not isinstance(obj, dict):
        return {}
    out: dict[str, float | None] = {}
    for k, v in obj.items():
        key = str(k)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            out[key] = None
        else:
            try:
                out[key] = float(v)
            except Exception:
                out[key] = None
    return out


def _compute_total_from_components(components: dict[str, float | None]) -> float:
    for k, v in components.items():
        if v is None:
            continue
        if _is_marks_col(k):
            return float(v)
    total = 0.0
    for v in components.values():
        if v is None:
            continue
        total += float(v)
    return total


def _backfill_components_for_file(session, file_id: str, faculty_user_id: str) -> None:
    """
    For files uploaded before we started storing per-component marks, reconstruct
    components from the source CSV/XLSX and persist into student_marks.components_json.

    Matching strategy:
    - Prefer roll_no (normalized)
    - Fallback to case-insensitive student_name
    """
    if not _student_marks_has_components(session):
        return

    rec = (
        session.query(UploadedFile)
        .filter(
            UploadedFile.id == file_id,
            UploadedFile.uploaded_by_user_id == faculty_user_id,
        )
        .first()
    )
    if not rec:
        return

    rows, headers = _parse_raw_rows(rec.file_path, rec.name)
    if not rows or not headers:
        return

    name_col = next((c for c in headers if _is_name_col(c)), None) or headers[0]
    roll_col = next((c for c in headers if _is_roll_col_high(c)), None) or next((c for c in headers if _is_roll_col(c)), None)
    excluded = {c for c in (name_col, roll_col) if c is not None}
    excluded |= {c for c in headers if _is_serial_like_col(c)}

    def _is_component_header(h: str) -> bool:
        clean = _clean_header(h)
        if not clean:
            return False
        if clean in {"grade", "result", "remarks", "remark", "status"}:
            return False
        tokens = {
            "assign", "assignment", "ass", "bit", "bits", "quiz", "mid", "midterm", "unit", "test",
            "internal", "external", "theory", "practical", "lab", "project", "viva", "descriptive",
        }
        return any(t in clean.replace(" ", "") for t in tokens)

    metric_cols = [c for c in headers if c not in excluded and (_col_numeric_count(c, rows) > 0 or _is_component_header(c))]
    if not metric_cols:
        return

    by_roll: dict[str, dict[str, float | None]] = {}
    by_name: dict[str, dict[str, float | None]] = {}

    for r in rows[:500]:
        raw_name = str(r.get(name_col, "")).strip()
        if not raw_name or raw_name.lower() in ("none", "nan", ""):
            continue
        raw_roll = _normalize_roll_no(r.get(roll_col, "")) if roll_col else ""

        comps: dict[str, float | None] = {}
        for c in metric_cols:
            comps[c] = _try_parse_number(r.get(c))

        if raw_roll:
            by_roll[raw_roll] = comps
        by_name[raw_name.strip().lower()] = comps

    if not by_roll and not by_name:
        return

    marks_rows = (
        session.query(StudentMark)
        .filter(StudentMark.uploaded_file_id == file_id)
        .all()
    )
    changed = 0
    for sm in marks_rows:
        if sm.components_json:
            continue
        roll = (sm.roll_no or "").strip()
        name_key = (sm.student_name or "").strip().lower()
        comps = by_roll.get(roll) if roll else None
        if comps is None:
            comps = by_name.get(name_key)
        if comps is None:
            continue
        sm.components_json = json.dumps(comps, ensure_ascii=False)
        changed += 1

    if changed:
        session.commit()


def get_uploaded_file_marks(file_id: str, faculty_user_id: str) -> dict:
    """Return persisted marks rows (including per-component scores) for an upload (ownership enforced)."""
    get_file_by_id(file_id, faculty_user_id)

    with _db.SessionLocal() as session:
        has_components = _student_marks_has_components(session)
        if not has_components:
            raise HTTPException(
                status_code=400,
                detail="Per-component marks require a DB migration. Run `python -m alembic -c database/alembic.ini upgrade head` and restart the backend.",
            )
        rows = (
            session.query(
                StudentMark.id,
                StudentMark.student_name,
                StudentMark.roll_no,
                StudentMark.marks,
                StudentMark.components_json if has_components else None,
            )
            .filter(StudentMark.uploaded_file_id == file_id)
            .order_by(StudentMark.roll_no.asc().nulls_last(), StudentMark.student_name.asc())
            .all()
        )

        parsed = []
        all_keys: set[str] = set()
        for (mid, student_name, roll_no, total, components_json) in rows:
            comps = _loads_components(components_json) if has_components else {}
            all_keys.update(comps.keys())
            parsed.append(((mid, student_name, roll_no, total), comps))

        # Hide accidental serial-number columns that may have been stored as components.
        def _looks_like_serial_component_key(key: str) -> bool:
            if _is_serial_like_col(key):
                return True
            values: list[int] = []
            for (_meta, comps) in parsed[:25]:
                v = comps.get(key)
                if v is None:
                    continue
                try:
                    n = float(v)
                except Exception:
                    continue
                if abs(n - round(n)) > 1e-9:
                    return False
                values.append(int(round(n)))
                if len(values) >= 12:
                    break
            if len(values) < 6:
                return False
            uniq = set(values)
            if len(uniq) / len(values) < 0.9:
                return False
            mn, mx = min(values), max(values)
            if mx > 200:
                return False
            return mn in (0, 1) and mx <= len(uniq) + 3

        serial_keys = {k for k in all_keys if _looks_like_serial_component_key(k)}
        if serial_keys:
            all_keys = {k for k in all_keys if k not in serial_keys}

        # Backfill legacy rows that were uploaded before per-component storage existed.
        if has_components and not all_keys:
            _backfill_components_for_file(session, file_id, faculty_user_id)
            rows2 = (
                session.query(StudentMark.id, StudentMark.student_name, StudentMark.roll_no, StudentMark.marks, StudentMark.components_json)
                .filter(StudentMark.uploaded_file_id == file_id)
                .order_by(StudentMark.roll_no.asc().nulls_last(), StudentMark.student_name.asc())
                .all()
            )
            parsed = []
            all_keys = set()
            for (mid, student_name, roll_no, total, components_json) in rows2:
                comps = _loads_components(components_json)
                all_keys.update(comps.keys())
                parsed.append(((mid, student_name, roll_no, total), comps))
            serial_keys = {k for k in all_keys if _looks_like_serial_component_key(k)}
            if serial_keys:
                all_keys = {k for k in all_keys if k not in serial_keys}

        # Preserve a sensible column order based on the source sheet headers when possible.
        columns_sorted = sorted(all_keys, key=lambda s: s.lower())
        columns = columns_sorted
        try:
            rec = (
                session.query(UploadedFile)
                .filter(
                    UploadedFile.id == file_id,
                    UploadedFile.uploaded_by_user_id == faculty_user_id,
                )
                .first()
            )
            if rec:
                raw_rows, headers = _parse_raw_rows(rec.file_path, rec.name)
                if raw_rows and headers:
                    name_col = next((c for c in headers if _is_name_col(c)), None) or headers[0]
                    roll_col = next((c for c in headers if _is_roll_col_high(c)), None) or next((c for c in headers if _is_roll_col(c)), None)
                    excluded = {c for c in (name_col, roll_col) if c is not None}
                    excluded |= {c for c in headers if _is_serial_like_col(c) or _looks_like_serial_col(c, raw_rows)}

                    def _is_component_header(h: str) -> bool:
                        clean = _clean_header(h)
                        if not clean:
                            return False
                        if clean in {"grade", "result", "remarks", "remark", "status"}:
                            return False
                        tokens = {
                            "assign", "assignment", "ass", "bit", "bits", "quiz", "mid", "midterm", "unit", "test",
                            "internal", "external", "theory", "practical", "lab", "project", "viva", "descriptive",
                        }
                        return any(t in clean.replace(" ", "") for t in tokens)

                    ordered = [
                        c for c in headers
                        if c not in excluded and (_col_numeric_count(c, raw_rows) > 0 or _is_component_header(c))
                    ]
                    ordered = [c for c in ordered if c in all_keys]
                    remaining = [c for c in columns_sorted if c not in set(ordered)]
                    columns = [*ordered, *remaining]
        except Exception:
            pass
        columns = _dedupe_total_like_columns(columns)
        return {
            "columns": columns,
            "rows": [
                {
                    "id": mid,
                    "name": (student_name or "").strip() or "Unknown",
                    "roll_no": (roll_no or "").strip(),
                    "total": float(total),
                    "components": {k: comps.get(k) for k in columns},
                }
                for ((mid, student_name, roll_no, total), comps) in parsed
            ],
        }


def update_uploaded_file_marks(file_id: str, faculty_user_id: str, marks: list[dict]) -> dict:
    """
    Update persisted marks rows for a specific uploaded file (ownership enforced).

    Expects a full list of rows with stable `id`s. Only the provided ids are updated.
    """
    get_file_by_id(file_id, faculty_user_id)

    if not marks:
        raise HTTPException(status_code=422, detail="No marks rows provided.")

    updates_by_id: dict[str, dict] = {}
    for m in marks:
        mid = (m.get("id") or "").strip()
        if not mid:
            raise HTTPException(status_code=422, detail="Each marks row must include an id.")
        updates_by_id[mid] = m

    with _db.SessionLocal() as session:
        has_components = _student_marks_has_components(session)
        existing = (
            session.query(StudentMark)
            .filter(
                StudentMark.uploaded_file_id == file_id,
                StudentMark.id.in_(list(updates_by_id.keys())),
            )
            .all()
        )
        if len(existing) != len(updates_by_id):
            found_ids = {r.id for r in existing}
            missing = [mid for mid in updates_by_id.keys() if mid not in found_ids]
            raise HTTPException(status_code=404, detail=f"Marks row(s) not found: {missing[:5]}")

        for row in existing:
            upd = updates_by_id[row.id]
            name = (upd.get("name") or "").strip()
            if not name:
                raise HTTPException(status_code=422, detail="Student name cannot be empty.")
            roll = _normalize_roll_no((upd.get("roll_no") or "").strip()) or None

            comps_in = upd.get("components") or {}
            if comps_in is not None and not isinstance(comps_in, dict):
                raise HTTPException(status_code=422, detail="components must be an object/dict.")

            comps: dict[str, float | None] = {}
            for k, v in (comps_in or {}).items():
                key = str(k)
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    comps[key] = None
                else:
                    try:
                        comps[key] = float(v)
                    except Exception:
                        raise HTTPException(status_code=422, detail=f"Invalid number for '{key}'.")

            if comps:
                if not has_components:
                    raise HTTPException(status_code=400, detail="Database not migrated for per-component marks. Run `alembic upgrade head` and re-upload the file.")
                total_val = _compute_total_from_components(comps)
            else:
                # Back-compat: accept "total" (preferred) or "marks" (older clients)
                raw_total = upd.get("total", upd.get("marks", row.marks))
                try:
                    total_val = float(raw_total)
                except Exception:
                    raise HTTPException(status_code=422, detail="Total must be a number.")

            row.student_name = name
            row.roll_no = roll
            row.marks = total_val
            if has_components and comps:
                row.components_json = json.dumps(comps, ensure_ascii=False)

        session.commit()

    return get_uploaded_file_marks(file_id, faculty_user_id)


def export_uploaded_file_marks_csv(file_id: str, faculty_user_id: str) -> tuple[str, str]:
    """Export the persisted (possibly edited) marks for a file as CSV."""
    with _db.SessionLocal() as session:
        rec = (
            session.query(UploadedFile)
            .filter(
                UploadedFile.id == file_id,
                UploadedFile.uploaded_by_user_id == faculty_user_id,
            )
            .first()
        )
        if not rec:
            raise HTTPException(status_code=404, detail=f"File '{file_id}' not found.")

        payload = get_uploaded_file_marks(file_id, faculty_user_id)
        columns: list[str] = payload.get("columns", [])
        rows: list[dict] = payload.get("rows", [])

        buf = io.StringIO()
        writer = csv.writer(buf)
        total_key = next((c for c in columns if _is_marks_col(c)), None)
        writer.writerow(["roll_no", "name", *columns] if total_key else ["roll_no", "name", *columns, "total"])
        for r in rows:
            comps = r.get("components") or {}
            out = [
                (r.get("roll_no") or ""),
                (r.get("name") or ""),
                *[(comps.get(c) if comps.get(c) is not None else "") for c in columns],
            ]
            if not total_key:
                out.append(float(r.get("total") or 0))
            writer.writerow(out)
        csv_text = buf.getvalue()

        base = (rec.name or "marks").rsplit(".", 1)[0]
        safe = "".join(ch if ch.isalnum() or ch in (" ", "_", "-", ".") else "_" for ch in base).strip() or "marks"
        filename = f"{safe}_edited.csv"
        return csv_text, filename
